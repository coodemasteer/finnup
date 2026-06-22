"""api/routers/templates.py — template download + Excel upload parsing"""
from __future__ import annotations

import io
import sys
import warnings
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

sys.path.insert(0, str(Path(__file__).parent.parent.parent))
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.models.lender_matcher import load_policies, compute_match_score, rank_lenders_meta

router = APIRouter()

# ── Cached lender policies (loaded once) ──────────────────────────────────────
_POLICIES: "pd.DataFrame | None" = None

def _get_policies() -> "pd.DataFrame":
    global _POLICIES
    if _POLICIES is None:
        try:
            _POLICIES = load_policies()
        except Exception:
            import pandas as _pd
            _POLICIES = _pd.DataFrame()
    return _POLICIES

# ── Column definitions ─────────────────────────────────────────────────────────
COLUMNS = [
    ("company_name",       "Company Name",                     "DEMO Corp",                     "Text (optional)"),
    ("product_name",       "Product Name",                     "Unsecured Business Loan",        "Unsecured Business Loan / Term Loan / Cash Credit/WCDL / LAP / Personal Loan / Housing Loan / Bill Discounting / Purchase Financing / Overdraft Facility"),
    ("location",           "Location (City)",                  "Mumbai",                        "Text — city name"),
    ("entity_type",        "Type of Entity",                   "Sole Proprietorship",            "Sole Proprietorship / Private Limited Company / Public Limited Company / Partnership / Limited Liability Partnership / Co-Operative / Individual"),
    ("loan_min_lakhs",     "Loan Amount Min (₹ Lakhs)",         8,                              "Numeric — in Lakhs (e.g. 8 = ₹8 Lakh)"),
    ("loan_max_lakhs",     "Loan Amount Max (₹ Lakhs)",        75,                              "Numeric — in Lakhs"),
    ("tenor_min_months",   "Tenor Min (months)",               12,                              "Numeric — e.g. 12 for 1 year"),
    ("tenor_max_months",   "Tenor Max (months)",               36,                              "Numeric — e.g. 36 for 3 years"),
    ("cibil_score",        "CIBIL Score",                      720,                             "Numeric — 300 to 900"),
    ("dpd90",              "DPD 90+ (last 12 months)",         0,                               "Numeric — count of DPD 90+ events"),
    ("overdue_accounts",   "Count of Overdue Accounts",        0,                               "Numeric — 0 or more"),
    ("overdue_amount",     "Total Overdue Amount (₹)",         0,                               "Numeric — rupees"),
    ("suit_filed",         "Suit Filed Count",                 0,                               "Numeric — 0 or more"),
    ("vintage_months",     "Business Vintage (months)",        18,                              "Numeric — months since incorporation (12–24)"),
    ("age_applicant",      "Age of Applicant (years)",         42,                              "Numeric — 21 to 70"),
    ("net_sales_lakhs",    "Net Sales (₹ Lakhs)",              7000,                            "Numeric — in Lakhs"),
    ("pat_lakhs",          "Profit After Tax (₹ Lakhs)",       210,                             "Numeric — in Lakhs (negative = loss)"),
    ("tnw_lakhs",          "Tangible Networth (₹ Lakhs)",      500,                             "Numeric — in Lakhs"),
    ("dscr",               "DSCR (Debt Service Coverage)",     1.2,                             "Numeric — ratio, e.g. 1.25"),
    ("current_ratio",      "Current Ratio",                    1.3,                             "Numeric — ratio, e.g. 1.3"),
    ("tol_tnw",            "TOL / TNW Ratio",                  1.5,                             "Numeric — ratio, e.g. 1.5"),
    ("inward_bounces",     "Inward Cheque Bounces",            0,                               "Numeric — count"),
    ("outward_bounces",    "Outward Cheque Bounces",           0,                               "Numeric — count"),
    ("enquiries_30d",      "Credit Enquiries (last 30 days)",  1,                               "Numeric — count"),
    ("new_sanctions_30d",  "New Sanctions (last 30 days)",     0,                               "Numeric — count"),
    ("gst_filing_3mo",     "GST Filing (last 3 months)",        "All filed",                     "All filed / 1 or 2 filed"),
    ("gst_filing_6mo",     "GST Filing (last 6 months)",        "All filed",                     "All filed / 1 or 2 not filed"),
    ("property_owned",     "Owned / Rented Property",          "Owned",                         "Owned / Rented"),
]

SAMPLE_ROW_2 = [
    "Alpha Traders", "Term Loan", "Delhi", "Private Limited",
    8, 75, 12, 36, 750, 0, 0, 0, 0, 20, 38, 7000, 210, 500, 1.5, 1.4, 1.2,
    0, 0, 0, 0, "All filed", "Partially filed", "Rented",
]

# Mapping from Excel display header → API column name used by the parsers
_DISPLAY_TO_API = {c[1]: c[0] for c in COLUMNS}


# ── Styling helpers ────────────────────────────────────────────────────────────
NAVY   = "1B3A6B"
TEAL   = "0D9488"
WHITE  = "FFFFFF"
LIGHT  = "F0F4F8"
AMBER  = "F59E0B"


def _make_workbook(multi_row: bool = False) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Borrowers" if multi_row else "Borrower Profile"

    header_fill   = PatternFill("solid", fgColor=NAVY)
    desc_fill     = PatternFill("solid", fgColor="E2E8F0")
    sample_fill   = PatternFill("solid", fgColor="F0FDFA")
    sample2_fill  = PatternFill("solid", fgColor="FFF7ED")
    bold_white    = Font(bold=True, color=WHITE, size=10)
    small_gray    = Font(color="475569", size=9, italic=True)
    normal        = Font(size=10)
    center        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin",   color="CBD5E1"),
        right=Side(style="thin",  color="CBD5E1"),
        top=Side(style="thin",    color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    headers = [c[1] for c in COLUMNS]
    sample  = [c[2] for c in COLUMNS]
    descs   = [c[3] for c in COLUMNS]

    # Row 1 — header
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill      = header_fill
        cell.font      = bold_white
        cell.alignment = center
        cell.border    = thin_border

    # Row 2 — description
    ws.append(descs)
    for cell in ws[2]:
        cell.fill      = desc_fill
        cell.font      = small_gray
        cell.alignment = left
        cell.border    = thin_border

    # Row 3 — sample data
    ws.append(sample)
    for cell in ws[3]:
        cell.fill      = sample_fill
        cell.font      = Font(color="0f766e", size=10)
        cell.alignment = left
        cell.border    = thin_border

    if multi_row:
        ws.append(SAMPLE_ROW_2)
        for cell in ws[4]:
            cell.fill      = sample2_fill
            cell.font      = Font(color="92400E", size=10)
            cell.alignment = left
            cell.border    = thin_border

    # Column widths
    col_widths = [20, 30, 18, 22, 22, 22, 14, 14, 20, 22, 22, 16, 24, 22, 22, 24, 24,
                  10, 15, 15, 22, 22, 24, 24, 25, 25, 22, 22]
    for i, w in enumerate(col_widths[:len(COLUMNS)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45
    ws.freeze_panes = "A3"

    return wb


# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.get("/template/single")
def download_single_template():
    """Download a pre-filled Excel template for a single borrower."""
    wb = _make_workbook(multi_row=False)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finnup_single_borrower_template.xlsx"},
    )


@router.get("/template/batch")
def download_batch_template():
    """Download a pre-filled Excel template for batch scoring."""
    wb = _make_workbook(multi_row=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finnup_batch_template.xlsx"},
    )


def _parse_df(contents: bytes) -> pd.DataFrame:
    """Read uploaded Excel/CSV, skip the description row (row index 1)."""
    buf = io.BytesIO(contents)
    try:
        df = pd.read_excel(buf, header=0)
    except Exception:
        buf.seek(0)
        df = pd.read_csv(buf, header=0)

    # Normalise display column names → API names so all downstream .get() calls work
    df = df.rename(columns=_DISPLAY_TO_API)

    # If 2nd row looks like a description row (non-numeric in numeric columns), drop it
    if not df.empty:
        first_val = str(df.iloc[0].get("cibil_score", df.iloc[0, 6] if len(df.columns) > 6 else ""))
        if not first_val.replace(".", "").replace("-", "").isnumeric():
            df = df.iloc[1:].reset_index(drop=True)
    return df


@router.post("/parse-single")
async def parse_single_excel(file: UploadFile = File(...)):
    """
    Parse the first data row of an uploaded Excel template and return
    a JSON object that maps directly to the LenderMatching form state.
    """
    contents = await file.read()
    try:
        df = _parse_df(contents)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read file: {exc}")

    if df.empty:
        raise HTTPException(status_code=400, detail="Uploaded file has no data rows.")

    r = df.iloc[0]

    def _get(col: str, default: Any) -> Any:
        val = r.get(col, default)
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return default
        return val

    # Normalize product_name to a valid dropdown option
    _RAW_PRODUCT = str(_get("product_name", "Unsecured Business Loan")).strip()
    _PRODUCT_NORM: dict[str, str] = {
        "bill discounting/ purchase financing": "Bill Discounting",
        "bill discounting/purchase financing":  "Bill Discounting",
        "bill discounting":                     "Bill Discounting",
        "purchase financing":                   "Purchase Financing",
        "purchase finance":                     "Purchase Financing",
        "lap":                                  "Loan Against Property",
        "loan against property":                "Loan Against Property",
        "cash credit/wcdl":                     "Cash Credit/WCDL",
        "cash credit":                          "Cash Credit/WCDL",
        "wcdl":                                 "Cash Credit/WCDL",
        "overdraft":                            "Overdraft Facility",
        "overdraft facility":                   "Overdraft Facility",
        "term loan":                            "Term Loan",
        "unsecured business loan":              "Unsecured Business Loan",
        "personal loan":                        "Personal Loan",
        "housing loan":                         "Housing Loan",
    }
    _product_name = _PRODUCT_NORM.get(_RAW_PRODUCT.lower(), _RAW_PRODUCT)

    # Normalize entity_type to canonical training value
    _ENTITY_NORM: dict[str, str] = {
        "llp":                           "Limited Liability Partnership",
        "limited liability partnership":  "Limited Liability Partnership",
        "private limited":               "Private Limited Company",
        "private limited company":       "Private Limited Company",
        "public limited":                "Public Limited Company",
        "public limited company":        "Public Limited Company",
        "sole proprietorship":           "Sole Proprietorship",
        "partnership":                   "Partnership",
        "co-operative":                  "Co-Operative",
        "cooperative":                   "Co-Operative",
        "individual":                    "Individual",
    }
    _raw_entity = str(_get("entity_type", "Sole Proprietorship")).strip()
    _entity_type = _ENTITY_NORM.get(_raw_entity.lower(), _raw_entity)

    # Normalize GST filing values
    _raw_gst3 = str(_get("gst_filing_3mo", "All filed")).strip()
    _gst3 = {"all filed": "All filed", "partially filed": "1 or 2 filed",
             "not filed": "1 or 2 filed", "1 or 2 filed": "1 or 2 filed"}.get(_raw_gst3.lower(), _raw_gst3)
    _raw_gst6 = str(_get("gst_filing_6mo", "All filed")).strip()
    _gst6 = {"all filed": "All filed", "partially filed": "1 or 2 not filed",
             "not filed": "1 or 2 not filed", "1 or 2 not filed": "1 or 2 not filed"}.get(_raw_gst6.lower(), _raw_gst6)

    try:
        result = {
            "company_name":   str(_get("company_name", "")),
            "entity_type":    _entity_type,
            "product_name":   _product_name,
            "location":       str(_get("location", "Mumbai")),
            "loan_min":       float(_get("loan_min_lakhs", 8)),
            "loan_max":       float(_get("loan_max_lakhs", 75)),
            "tenor_min":      int(float(_get("tenor_min_months", 12))),
            "tenor_max":      int(float(_get("tenor_max_months", 36))),
            "cibil":          int(float(_get("cibil_score", 720))),
            "dpd90":          int(float(_get("dpd90", 0))),
            "overdue_count":  int(float(_get("overdue_accounts", 0))),
            "overdue_amount": float(_get("overdue_amount", 0)),
            "suit_filed":     int(float(_get("suit_filed", 0))),
            "vintage":        int(float(_get("vintage_months", 18))),
            "age_app":        int(float(_get("age_applicant", 42))),
            "net_sales":      float(_get("net_sales_lakhs", 7000)),
            "pat":            float(_get("pat_lakhs", 210)),
            "tnw":            float(_get("tnw_lakhs", 500)),
            "dscr":           float(_get("dscr", 1.2)),
            "current_ratio":  float(_get("current_ratio", 1.3)),
            "tol_tnw":        float(_get("tol_tnw", 1.5)),
            "inward_bounces": int(float(_get("inward_bounces", 0))),
            "outward_bounces":int(float(_get("outward_bounces", 0))),
            "enq30":          int(float(_get("enquiries_30d", 1))),
            "ns30":           int(float(_get("new_sanctions_30d", 0))),
            "gst3":           _gst3,
            "gst6":           _gst6,
            "owned":          str(_get("property_owned", "Owned")),
        }
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Failed to parse row: {exc}")

    return result


@router.post("/batch-score-upload")
async def batch_score_from_upload(file: UploadFile = File(...)):
    """
    Score an uploaded Excel/CSV batch file using the trained model.
    The file must follow the finnup_batch_template.xlsx column layout.
    """
    from api.routers.batch import MODELS_PATH  # local import to avoid circular

    if not MODELS_PATH.exists():
        raise HTTPException(status_code=503, detail="No trained model found. Run /api/train first.")

    import pickle
    from src.features.engineering import engineer_features

    contents = await file.read()
    try:
        df_raw = _parse_df(contents)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read file: {exc}")

    if df_raw.empty:
        raise HTTPException(status_code=400, detail="Uploaded file has no data rows.")

    df = df_raw.copy()

    # _parse_df already applied _DISPLAY_TO_API so columns are now API names.
    # Rename API names → raw column names expected by engineer_features.
    api_to_raw = {
        "entity_type":      "Type of Entity",
        "cibil_score":      "CIBIL Score",
        "dpd90":            "cnt_dpd_90plus_last_12mo",
        "overdue_accounts": "Count of Overdue Accounts",
        "overdue_amount":   "Total Overdue Amount",
        "suit_filed":       "Suit Filed Count of Loans",
        "vintage_months":   "Vintage (in months)",
        "age_applicant":    "Age of applicant",
        "dscr":             "DSCR (Avg/Min)",
        "current_ratio":    "Current Ratio",
        "tol_tnw":          "TOL/ TNW",
        "inward_bounces":   "Total Number of Inward cheque bounces",
        "outward_bounces":  "Total Number of Outward cheque bounces",
        "enquiries_30d":    "enquiry_last30days",
        "new_sanctions_30d": "New sanction in the last 30 days",
        "gst_filing_3mo":   "GST Filing in the past 3 months",
        "gst_filing_6mo":   "GST Filing in the past 6 months",
        "property_owned":   "Owned/Rented Property",
        "tenor_min_months": "Tenor Min",
        "tenor_max_months": "Tenor Max",
    }
    df = df.rename(columns=api_to_raw)

    # Convert Lakh columns → ₹ (columns are now API names after _parse_df)
    for lakh_col, raw_col in [
        ("loan_min_lakhs",  "Loan Amount Min"),
        ("loan_max_lakhs",  "Loan Amount Max"),
        ("net_sales_lakhs", "Net Sales"),
        ("pat_lakhs",       "Profit After Tax"),
        ("tnw_lakhs",       "Tangible Networth (TNW)"),
    ]:
        if lakh_col in df.columns:
            df[raw_col] = pd.to_numeric(df[lakh_col], errors="coerce").fillna(0) * 100_000
            df.drop(columns=[lakh_col], inplace=True)

    # Add placeholder columns required by engineer_features
    for col, default in [
        ("Loan Amount Min", 2_000_000),
        ("Loan Amount Max", 5_000_000),
        ("Tenor Min", 1), ("Tenor Max", 3),
        ("Rate of interest Min", 11), ("Rate of interest Max", 18),
        ("Total number of active accounts", 0),
        ("New sanction in the last 90 days", 0),
        ("enquiry_last7days", 0),
        ("cnt_dpd_0plus_last_12mo", 0),
        ("Industry", 6),
        ("Total Amount of Credit Transactions", 0),
        ("Total Amount of Debit Transactions", 0),
        ("Average EOD Balance", 0),
    ]:
        if col not in df.columns:
            df[col] = default

    if "company_name" not in df.columns:
        df["company_name"] = [f"Borrower_{i+1}" for i in range(len(df))]

    df["loan_approved"] = 0  # placeholder

    with open(MODELS_PATH, "rb") as f:
        saved = pickle.load(f)
    models    = saved["models"]
    feat_names = saved["features"]

    X = engineer_features(df)
    for c in feat_names:
        if c not in X.columns:
            X[c] = 0
    X = X[feat_names].fillna(0)

    best_mdl = models.get("best") or list(models.values())[0]
    proba    = best_mdl.predict_proba(X)[:, 1]

    # ── Engine 2: lender policy matching per row ───────────────────────────
    policies = _get_policies()

    def _top_lenders(borrower_row: pd.Series, p: float) -> list[dict]:
        """Return top-3 lender matches for one borrower row."""
        if policies.empty:
            return []
        try:
            match_df = compute_match_score(borrower_row, policies)
            ranked   = rank_lenders_meta(p, match_df, top_n=3)
            return [
                {
                    "lender_name":    str(r["lender_name"]),
                    "match_score":    round(float(r["match_score"]) * 100, 1),
                    "combined_score": round(float(r["combined_score"]) * 100, 1),
                }
                for _, r in ranked.iterrows()
            ]
        except Exception:
            return []

    def _risk_band(p: float) -> str:
        if p >= 0.70: return "Very High"
        if p >= 0.50: return "High"
        if p >= 0.30: return "Medium"
        return "Low"

    def _sv(row: pd.Series, col: str, default: str = "") -> str:
        v = row.get(col)
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        return str(v)

    def _fv(row: pd.Series, col: str, default: float = 0.0) -> float:
        v = row.get(col)
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        try:
            return float(v)
        except Exception:
            return default

    def _iv(row: pd.Series, col: str, default: int = 0) -> int:
        return int(_fv(row, col, float(default)))

    rows = []
    for i, (_, row) in enumerate(df.iterrows()):
        p = float(proba[i])
        rows.append({
            "index":            i + 1,
            # ── Identity ──────────────────────────────────────────────────
            "company_name":     _sv(row, "company_name", f"Borrower_{i+1}"),
            "product_name":     _sv(row, "product_name"),
            "location":         _sv(row, "location"),
            "entity_type":      _sv(row, "Type of Entity"),
            # ── Loan parameters (stored in Lakhs for form parity) ─────────
            "loan_min":         round(_fv(row, "Loan Amount Min") / 100_000, 2),
            "loan_max":         round(_fv(row, "Loan Amount Max") / 100_000, 2),
            "tenor_min":        _iv(row, "Tenor Min", 12),
            "tenor_max":        _iv(row, "Tenor Max", 36),
            # ── Credit history ────────────────────────────────────────────
            "cibil_score":      _fv(row, "CIBIL Score"),
            "dpd90":            _iv(row, "cnt_dpd_90plus_last_12mo"),
            "overdue_accounts": _iv(row, "Count of Overdue Accounts"),
            "overdue_amount":   _fv(row, "Total Overdue Amount"),
            "suit_filed":       _iv(row, "Suit Filed Count of Loans"),
            # ── Business profile ──────────────────────────────────────────
            "vintage":          _iv(row, "Vintage (in months)"),
            "age_app":          _iv(row, "Age of applicant", 42),
            # ── Financials (stored in Lakhs for form parity) ──────────────
            "net_sales":        round(_fv(row, "Net Sales") / 100_000, 2),
            "pat":              round(_fv(row, "Profit After Tax") / 100_000, 2),
            "tnw":              round(_fv(row, "Tangible Networth (TNW)") / 100_000, 2),
            # ── Ratios ────────────────────────────────────────────────────
            "dscr":             round(_fv(row, "DSCR (Avg/Min)", 1.2), 2),
            "current_ratio":    round(_fv(row, "Current Ratio", 1.3), 2),
            "tol_tnw":          round(_fv(row, "TOL/ TNW", 1.5), 1),
            # ── Banking & GST ─────────────────────────────────────────────
            "inward_bounces":   _iv(row, "Total Number of Inward cheque bounces"),
            "outward_bounces":  _iv(row, "Total Number of Outward cheque bounces"),
            "enq30":            _iv(row, "enquiry_last30days"),
            "ns30":             _iv(row, "New sanction in the last 30 days"),
            "gst3":             _sv(row, "GST Filing in the past 3 months", "All filed"),
            "gst6":             _sv(row, "GST Filing in the past 6 months", "All filed"),
            "owned":            _sv(row, "Owned/Rented Property", "Owned"),
            # ── Scores ────────────────────────────────────────────────────
            "p_approved_pct":   round(p * 100, 1),
            "label":            "Uploaded",
            "risk_band":        _risk_band(p),
            "top_lenders":      _top_lenders(row, p),
        })

    histogram = [float(p) for p in proba]

    high_prob   = int(sum(1 for p in proba if p >= 0.50))
    medium_prob = int(sum(1 for p in proba if 0.30 <= p < 0.50))
    low_prob    = int(sum(1 for p in proba if p < 0.30))

    return {
        "total":              len(rows),
        "high_prob":          high_prob,
        "medium_prob":        medium_prob,
        "low_prob":           low_prob,
        "confirmed_approved": 0,
        "confirmed_rejected": 0,
        "rows":               rows,
        "histogram":          histogram,
        "source":             "upload",
    }
